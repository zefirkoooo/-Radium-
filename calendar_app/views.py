import os
import logging
from datetime import datetime, timedelta
from io import BytesIO
import glob
import signal
import shutil

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import login, authenticate, logout
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from .models import CustomUser, Meal, DayMenu, UserSelection, FoodCategory
from .forms import UserRegistrationForm, MealUploadForm, UserSelectionForm
import pandas as pd
from django.utils import timezone
from django.contrib.auth.forms import PasswordChangeForm, SetPasswordForm
from django.contrib.admin.views.decorators import staff_member_required
from openpyxl import load_workbook
from django.conf import settings
from django.db.models import Count, Q, Prefetch
import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from copy import copy
import sys
import traceback
from django.views.decorators.http import require_http_methods, require_POST
import json
import tempfile
from django.db import transaction

from django.http import JsonResponse
from django.db.models import Q
import json

# Настройка логирования с правильной кодировкой
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)  # Reduced from DEBUG

# Создаем обработчик для вывода в файл с указанием кодировки
file_handler = logging.FileHandler(os.path.join(settings.BASE_DIR, 'logs', 'upload_debug.log'), encoding='utf-8')
file_handler.setLevel(logging.INFO)

# Создаем обработчик для вывода в консоль
console_handler = logging.StreamHandler(sys.stdout)
console_handler.setLevel(logging.WARNING)

# Создаем форматтер
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
console_handler.setFormatter(formatter)

# Добавляем обработчики к логгеру
logger.addHandler(file_handler)
logger.addHandler(console_handler)


def search_users_api(request):
    """API для поиска пользователей с автоподсказками"""
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    query = request.GET.get('q', '').strip()

    if len(query) < 2:
        return JsonResponse({'users': []})

    try:
        # Поиск пользователей по имени (регистронезависимый)
        users = CustomUser.objects.filter(
            Q(username__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query)
        ).values('id', 'username', 'first_name', 'last_name')[:10]

        # Форматируем результаты
        user_list = []
        for user in users:
            display_name = user['username']
            if user['first_name'] or user['last_name']:
                full_name = f"{user['first_name']} {user['last_name']}".strip()
                display_name = f"{user['username']} ({full_name})"

            user_list.append({
                'id': user['id'],
                'username': user['username'],
                'display_name': display_name
            })

        return JsonResponse({'users': user_list})

    except Exception as e:
        logger.error(f'Error searching users: {str(e)}')
        return JsonResponse({'error': 'Ошибка поиска'}, status=500)


# Замените функцию user_calendar_api в views.py

def user_calendar_api(request):
    """API для получения календаря питания пользователя - только текущая неделя"""
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    username = request.GET.get('username', '').strip()

    if not username:
        return JsonResponse({'error': 'Имя пользователя не указано'}, status=400)

    try:
        # Найти пользователя
        try:
            user = CustomUser.objects.get(username=username)
        except CustomUser.DoesNotExist:
            return JsonResponse({'error': 'Пользователь не найден'}, status=404)

        # Получить даты только текущей недели
        current_date = timezone.now().date()
        current_week_start = current_date - timedelta(days=current_date.weekday())

        # Получить меню и выборы пользователя только для текущей недели
        calendar_data = {
            'current_week': get_week_data(user, current_week_start)
        }

        return JsonResponse({
            'success': True,
            'username': username,
            'calendar': calendar_data
        })

    except Exception as e:
        logger.error(f'Error fetching user calendar: {str(e)}')
        return JsonResponse({'error': 'Ошибка при загрузке календаря'}, status=500)


def get_week_data(user, week_start):
    """Получить данные о неделе для пользователя"""
    week_data = []

    for day_offset in range(5):  # Понедельник - Пятница
        date = week_start + timedelta(days=day_offset)

        # Получить меню на этот день
        try:
            day_menu = DayMenu.objects.get(date=date)
        except DayMenu.DoesNotExist:
            # Если меню нет, добавляем день без выборов
            day_data = {
                'date': date.strftime('%d.%m.%Y'),
                'day_name': ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница'][day_offset],
                'has_menu': False,
                'selections': {
                    'not_eating': False,
                    'salad': None,
                    'soup': None,
                    'main': None,
                    'side': None,
                    'bakery': None,
                }
            }
            week_data.append(day_data)
            continue

        # Получить выбор пользователя
        try:
            selection = UserSelection.objects.select_related(
                'selected_salad', 'selected_soup', 'selected_main',
                'selected_side', 'selected_bakery'
            ).get(user=user, day_menu=day_menu)
        except UserSelection.DoesNotExist:
            selection = None

        # Форматировать данные дня
        day_data = {
            'date': date.strftime('%d.%m.%Y'),
            'day_name': day_menu.get_day_display(),
            'has_menu': True,
            'selections': {
                'not_eating': selection.not_eating if selection else False,
                'salad': selection.selected_salad.name if selection and selection.selected_salad else None,
                'soup': selection.selected_soup.name if selection and selection.selected_soup else None,
                'main': selection.selected_main.name if selection and selection.selected_main else None,
                'side': selection.selected_side.name if selection and selection.selected_side else None,
                'bakery': selection.selected_bakery.name if selection and selection.selected_bakery else None,
            }
        }

        week_data.append(day_data)

    return week_data


def get_week_data(user, week_start):
    """Получить данные о неделе для пользователя"""
    week_data = []

    for day_offset in range(5):  # Понедельник - Пятница
        date = week_start + timedelta(days=day_offset)

        # Получить меню на этот день
        try:
            day_menu = DayMenu.objects.get(date=date)
        except DayMenu.DoesNotExist:
            continue

        # Получить выбор пользователя
        try:
            selection = UserSelection.objects.get(user=user, day_menu=day_menu)
        except UserSelection.DoesNotExist:
            selection = None

        # Форматировать данные дня
        day_data = {
            'date': date.strftime('%d.%m.%Y'),
            'day_name': day_menu.get_day_display(),
            'selections': {
                'not_eating': selection.not_eating if selection else False,
                'salad': selection.selected_salad.name if selection and selection.selected_salad else None,
                'soup': selection.selected_soup.name if selection and selection.selected_soup else None,
                'main': selection.selected_main.name if selection and selection.selected_main else None,
                'side': selection.selected_side.name if selection and selection.selected_side else None,
                'bakery': selection.selected_bakery.name if selection and selection.selected_bakery else None,
            }
        }

        week_data.append(day_data)

    return week_data


def cleanup_old_files():
    """Удаляет старые Excel файлы, оставляя только 8 самых новых"""
    try:
        # Получаем список всех Excel файлов в директории
        excel_files = []
        for file in os.listdir(settings.MEDIA_ROOT):
            if file.endswith('.xlsx'):
                file_path = os.path.join(settings.MEDIA_ROOT, file)
                excel_files.append((file_path, os.path.getctime(file_path)))

        # Сортируем файлы по дате создания (от новых к старым)
        excel_files.sort(key=lambda x: x[1], reverse=True)

        # Оставляем только 8 самых новых файлов
        files_to_delete = excel_files[8:]

        # Удаляем старые файлы
        for file_path, _ in files_to_delete:
            try:
                os.remove(file_path)
                logger.info(f"Удален старый файл: {file_path}")
            except Exception as e:
                logger.error(f"Ошибка при удалении файла {file_path}: {str(e)}")

    except Exception as e:
        logger.error(f"Ошибка при очистке старых файлов: {str(e)}")


def login_view(request):
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        if not username or not password:
            messages.error(request, 'Пожалуйста, введите имя пользователя и пароль')
            return render(request, 'calendar_app/login.html')

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            next_url = request.GET.get('next', 'home')
            messages.success(request, f'Добро пожаловать, {user.username}!')
            return redirect(next_url)
        else:
            messages.error(request, 'Неверное имя пользователя или пароль')

    return render(request, 'calendar_app/login.html')


def is_admin(user):
    if not user.is_authenticated:
        return False
    return user.is_admin or user.is_superuser or user.is_staff


@login_required
def home(request):
    logger = logging.getLogger('calendar_app')

    # Get current and next week dates
    current_date = timezone.now().date()
    current_week_start = current_date - timedelta(days=current_date.weekday())
    next_week_start = current_week_start + timedelta(days=7)

    try:
        # Оптимизированные запросы с prefetch_related и select_related
        current_week_menu = DayMenu.objects.filter(
            date__range=[current_week_start, current_week_start + timedelta(days=4)]
        ).order_by('date').prefetch_related(
            'salads', 'soups', 'main_courses', 'sides', 'bakery'
        )

        next_week_menu = DayMenu.objects.filter(
            date__range=[next_week_start, next_week_start + timedelta(days=4)]
        ).order_by('date').prefetch_related(
            'salads', 'soups', 'main_courses', 'sides', 'bakery'
        )

        # Оптимизированный запрос выборов пользователя
        user_selections = UserSelection.objects.filter(
            user=request.user,
            day_menu__date__range=[current_week_start, next_week_start + timedelta(days=4)]
        ).select_related(
            'day_menu', 'selected_salad', 'selected_soup', 'selected_main',
            'selected_side', 'selected_bakery'
        )

        # Create a mapping of day_menu_id to user_selection
        selection_map = {sel.day_menu_id: sel for sel in user_selections}

        # Attach user selections to menus
        for menu in current_week_menu:
            menu.user_selection = selection_map.get(menu.id)

        for menu in next_week_menu:
            menu.user_selection = selection_map.get(menu.id)

        context = {
            'current_week_menu': current_week_menu,
            'next_week_menu': next_week_menu,
            'is_admin': is_admin(request.user),
            'current_week_start': current_week_start,
            'next_week_start': next_week_start,
        }

        # Handle file upload
        if request.method == 'POST' and 'excel_file' in request.FILES:
            if not (request.user.is_admin or request.user.is_superuser):
                messages.error(request, "У вас нет прав для загрузки файлов")
                return render(request, 'calendar_app/home.html', context)

            excel_file = request.FILES['excel_file']
            logger.info(f"Processing file upload: {excel_file.name} ({excel_file.size} bytes)")

            # Validate file type
            if not excel_file.name.lower().endswith(('.xlsx', '.xls')):
                messages.error(request, "Файл должен быть в формате Excel (.xlsx или .xls)")
                return render(request, 'calendar_app/home.html', context)

            # Validate file size (10MB limit)
            if excel_file.size > 10 * 1024 * 1024:
                messages.error(request, "Размер файла не должен превышать 10MB")
                return render(request, 'calendar_app/home.html', context)

            try:
                # --- ОГРАНИЧЕНИЕ НА 5 ФАЙЛОВ ---
                excel_files = sorted(
                    glob.glob(os.path.join(settings.MEDIA_ROOT, 'menu_*.xlsx')),
                    key=os.path.getmtime
                )
                while len(excel_files) >= 5:
                    try:
                        os.remove(excel_files[0])
                        logger.info(f"Удалён старый файл: {excel_files[0]}")
                    except OSError as e:
                        logger.warning(f"Не удалось удалить файл {excel_files[0]}: {e}")
                    excel_files.pop(0)
                # --- КОНЕЦ ОГРАНИЧЕНИЯ ---

                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                file_path = os.path.join(settings.MEDIA_ROOT, f'menu_{timestamp}.xlsx')

                # Улучшенное сохранение файла
                try:
                    with open(file_path, 'wb+') as destination:
                        for chunk in excel_file.chunks():
                            destination.write(chunk)
                    logger.info(f"File saved to: {file_path}")
                except IOError as e:
                    logger.error(f"Ошибка сохранения файла: {e}")
                    messages.error(request, "Ошибка при сохранении файла")
                    return render(request, 'calendar_app/home.html', context)

                # Очищаем старые файлы после успешной загрузки
                cleanup_old_files()

                # Получаем все блюда, которые используются в текущем и следующем меню (оптимизированный запрос)
                current_week_end = current_week_start + timedelta(days=4)
                next_week_end = next_week_start + timedelta(days=4)

                used_meals = Meal.objects.filter(
                    Q(day_menus_as_salad__date__range=[current_week_start, current_week_end]) |
                    Q(day_menus_as_soup__date__range=[current_week_start, current_week_end]) |
                    Q(day_menus_as_main__date__range=[current_week_start, current_week_end]) |
                    Q(day_menus_as_side__date__range=[current_week_start, current_week_end]) |
                    Q(day_menus_as_bakery__date__range=[current_week_start, current_week_end]) |
                    Q(day_menus_as_salad__date__range=[next_week_start, next_week_end]) |
                    Q(day_menus_as_soup__date__range=[next_week_start, next_week_end]) |
                    Q(day_menus_as_main__date__range=[next_week_start, next_week_end]) |
                    Q(day_menus_as_side__date__range=[next_week_start, next_week_end]) |
                    Q(day_menus_as_bakery__date__range=[next_week_start, next_week_end])
                ).distinct()

                # Удаляем только те блюда, которые не используются в меню
                Meal.objects.exclude(id__in=used_meals.values_list('id', flat=True)).delete()
                logger.info("Cleared unused dishes before importing new ones")

                # Move next week's menus to current week if they exist
                next_week_menus = DayMenu.objects.filter(
                    date__range=[next_week_start, next_week_start + timedelta(days=4)]
                ).order_by('date').prefetch_related('userselection_set')

                if next_week_menus.exists():
                    logger.info("Found existing menus for next week, moving them to current week")

                    # Use transaction for atomicity
                    with transaction.atomic():
                        # First, delete current week's menus
                        DayMenu.objects.filter(
                            date__range=[current_week_start, current_week_start + timedelta(days=4)]
                        ).delete()
                        logger.info("Deleted current week's menus")

                        # Move next week's menus to current week
                        for day_offset in range(5):  # Monday to Friday
                            try:
                                menu = next_week_menus[day_offset]
                                old_date = menu.date
                                new_date = current_week_start + timedelta(days=day_offset)

                                # Create new menu for current week
                                new_menu = DayMenu.objects.create(date=new_date)

                                # Copy all relations
                                new_menu.salads.set(menu.salads.all())
                                new_menu.soups.set(menu.soups.all())
                                new_menu.main_courses.set(menu.main_courses.all())
                                new_menu.sides.set(menu.sides.all())
                                new_menu.bakery.set(menu.bakery.all())

                                # Copy user selections
                                for selection in menu.userselection_set.all():
                                    UserSelection.objects.create(
                                        user=selection.user,
                                        day_menu=new_menu,
                                        selected_salad=selection.selected_salad,
                                        selected_soup=selection.selected_soup,
                                        selected_main=selection.selected_main,
                                        selected_side=selection.selected_side,
                                        selected_bakery=selection.selected_bakery,
                                        not_eating=selection.not_eating
                                    )

                                logger.info(f"Copied menu from {old_date} to {new_date} with all selections")
                            except IndexError:
                                logger.warning(f"No menu found for day offset {day_offset}")
                            except Exception as e:
                                logger.error(f"Error copying menu: {str(e)}")

                        # Delete any remaining next week's menus
                        DayMenu.objects.filter(
                            date__range=[next_week_start, next_week_start + timedelta(days=4)]
                        ).delete()
                        logger.info("Deleted any remaining next week's menus")

                # Choose parser based on the submitted value
                parser_type = request.POST.get('parser_type', 'standard')
                logger.info(f"Using {parser_type} parser")

                if parser_type == 'smart':
                    # Use smart parser
                    parse_excel_smart(file_path, next_week_start)
                    messages.success(request, "Меню успешно загружено с использованием умного парсера")
                else:
                    # Use standard parser
                    wb = load_workbook(file_path, data_only=True)
                    ws = wb.active
                    logger.info(f"Loaded Excel file, max row: {ws.max_row}, max column: {ws.max_column}")

                    # Map days to Excel columns
                    day_columns = {
                        0: 'B',  # Понедельник
                        1: 'D',  # Вторник
                        2: 'F',  # Среда
                        3: 'H',  # Четверг
                        4: 'J'  # Пятница
                    }

                    # Category mappings (name to model field)
                    category_mappings = {
                        'салаты': ('Салаты', 'salads'),
                        'супы': ('Супы', 'soups'),
                        'горячие блюда': ('Горячие блюда', 'main_courses'),
                        'горячее': ('Горячие блюда', 'main_courses'),
                        'гарниры': ('Гарниры', 'sides'),
                        'выпечка': ('Выпечка', 'bakery')
                    }

                    # First, verify the structure
                    logger.info("Verifying Excel structure...")
                    for col in day_columns.values():
                        day_cell = ws[f'{col}2']
                        if not day_cell.value:
                            logger.warning(f"No day found in cell {col}2")

                    # Create food categories if they don't exist
                    unique_categories = set(cat[0] for cat in category_mappings.values())
                    for category_name in unique_categories:
                        cat, created = FoodCategory.objects.get_or_create(name=category_name)
                        logger.info(f"Category {category_name}: {'created' if created else 'already exists'}")

                    # Read category positions from column A
                    category_positions = {}
                    current_category = None
                    category_start = None

                    # Scan column A for categories and their positions
                    for row in range(3, ws.max_row + 1):
                        cell_value = ws[f'A{row}'].value
                        if cell_value and isinstance(cell_value, str):
                            cell_value = cell_value.lower().strip()

                            # Check if this is a category
                            for cat_key in category_mappings.keys():
                                if cat_key in cell_value:
                                    if current_category and category_start:
                                        category_positions[current_category] = (category_start, row - 1)
                                    current_category = category_mappings[cat_key][0]
                                    category_start = row
                                    break

                    # Add the last category range
                    if current_category and category_start:
                        category_positions[current_category] = (category_start, ws.max_row)

                    logger.info("Found category positions:")
                    for category, (start, end) in category_positions.items():
                        logger.info(f"{category}: rows {start}-{end}")

                    # Process each day of the week
                    for day_offset, col in day_columns.items():
                        date = next_week_start + timedelta(days=day_offset)
                        menu = DayMenu.objects.create(date=date)
                        logger.info(f"Created menu for {date}")

                        # Process each category
                        for category_name, positions in category_positions.items():
                            start_row, end_row = positions
                            field_name = next(field for cat, field in category_mappings.values()
                                              if cat == category_name)

                            # Get or create meals in this category
                            for row in range(start_row, end_row + 1):
                                cell_value = ws[f'{col}{row}'].value
                                if cell_value:
                                    name, description = parse_meal_name(str(cell_value))
                                    if name:
                                        try:
                                            # Create new meal
                                            category = FoodCategory.objects.get(name=category_name)
                                            meal = Meal.objects.create(
                                                name=name,
                                                description=description,
                                                category=category,
                                                excel_row=row
                                            )
                                            logger.info(f"Created new meal: {name} at row {row}")

                                            # Add meal to the appropriate menu field
                                            getattr(menu, field_name).add(meal)
                                            logger.info(f"Added {name} to {field_name} for {date}")
                                        except Exception as e:
                                            logger.error(f"Error creating meal '{name}': {str(e)}")

                    wb.close()
                    messages.success(request, "Меню успешно загружено")
                    logger.info("Menu import completed successfully")

                    # Redirect to manage_dishes page after successful upload
                    return redirect('manage_dishes')

                return redirect('home')

            except Exception as e:
                logger.error(f"Error handling file: {str(e)}", exc_info=True)
                messages.error(request, f"Ошибка при обработке файла: {str(e)}")
                return render(request, 'calendar_app/home.html', context)

        return render(request, 'calendar_app/home.html', context)

    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}", exc_info=True)
        messages.error(request, "Произошла непредвиденная ошибка")
        return render(request, 'calendar_app/home.html', {
            'current_week_menu': [],
            'next_week_menu': [],
            'is_admin': is_admin(request.user),
            'current_week_start': current_week_start,
            'next_week_start': next_week_start,
        })


@login_required
def day_detail(request, day_id):
    try:
        # Оптимизированный запрос с prefetch_related
        day_menu = get_object_or_404(
            DayMenu.objects.prefetch_related(
                'salads', 'soups', 'main_courses', 'sides', 'bakery'
            ),
            id=day_id
        )

        user_selection = UserSelection.objects.filter(
            user=request.user,
            day_menu=day_menu
        ).select_related(
            'selected_salad', 'selected_soup', 'selected_main',
            'selected_side', 'selected_bakery'
        ).first()

        current_date = timezone.now().date()
        current_week_start = current_date - timedelta(days=current_date.weekday())
        next_week_start = current_week_start + timedelta(days=7)

        if request.method == 'POST':
            # Используем транзакцию для атомарности
            with transaction.atomic():
                form = UserSelectionForm(request.POST, instance=user_selection)
                if form.is_valid():
                    selection = form.save(commit=False)
                    selection.user = request.user
                    selection.day_menu = day_menu

                    # Проверяем конфликты между полноценным блюдом и гарниром
                    if selection.selected_main and selection.selected_main.is_complete_dish and selection.selected_side:
                        form.add_error('selected_side', 'Нельзя выбрать гарнир к полноценному блюду')
                        return render(request, 'calendar_app/day_detail.html', {
                            'day_menu': day_menu,
                            'form': form,
                            'categories': [
                                ('Салаты', 'fas fa-leaf', '🥗', 'selected_salad', day_menu.salads.all()),
                                ('Супы', 'fas fa-soup', '🍲', 'selected_soup', day_menu.soups.all()),
                                ('Горячие блюда', 'fas fa-drumstick-bite', '🍖', 'selected_main',
                                 day_menu.main_courses.all()),
                                ('Гарниры', 'fas fa-carrot', '🥘', 'selected_side', day_menu.sides.all()),
                                ('Выпечка', 'fas fa-bread-slice', '🥨', 'selected_bakery', day_menu.bakery.all()),
                            ],
                            'next_week_start': next_week_start,
                        })

                    # Если выбрана опция "НЕ ЕМ", очищаем все выбранные блюда
                    if selection.not_eating:
                        selection.selected_salad = None
                        selection.selected_soup = None
                        selection.selected_main = None
                        selection.selected_side = None
                        selection.selected_bakery = None

                    selection.save()
                    messages.success(request, 'Ваш выбор успешно сохранен!')

                    if 'save_and_next' in request.POST:
                        next_day = DayMenu.objects.filter(
                            date__gt=day_menu.date
                        ).order_by('date').first()

                        if next_day:
                            return redirect('day_detail', day_id=next_day.id)
                        else:
                            messages.info(request, 'Вы достигли конца недели')
                            return redirect('home')

                    return redirect('home')
        else:
            form = UserSelectionForm(instance=user_selection)

        categories = [
            ('Салаты', 'fas fa-leaf', '🥗', 'selected_salad', day_menu.salads.all()),
            ('Супы', 'fas fa-temperature-high', '🍲', 'selected_soup', day_menu.soups.all()),
            ('Горячие блюда', 'fas fa-drumstick-bite', '🍖', 'selected_main', day_menu.main_courses.all()),
            ('Гарниры', 'fas fa-carrot', '🥘', 'selected_side', day_menu.sides.all()),
            ('Выпечка', 'fas fa-bread-slice', '🥨', 'selected_bakery', day_menu.bakery.all()),
        ]

        context = {
            'day_menu': day_menu,
            'form': form,
            'categories': categories,
            'next_week_start': next_week_start,
        }

        return render(request, 'calendar_app/day_detail.html', context)
    except DayMenu.DoesNotExist:
        messages.error(request, 'Меню не найдено')
        return redirect('home')
    except Exception as e:
        messages.error(request, f'Произошла ошибка: {str(e)}')
        return redirect('home')


@user_passes_test(is_admin)
def user_management(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        action = request.POST.get('action')

        # Используем select_for_update для избежания race conditions
        with transaction.atomic():
            user = get_object_or_404(
                CustomUser.objects.select_for_update(),
                id=user_id
            )

            if action == 'make_admin':
                user.is_admin = True
                user.save(update_fields=['is_admin'])
                messages.success(request, f'Пользователь {user.username} теперь администратор')
            elif action == 'remove_admin':
                user.is_admin = False
                user.save(update_fields=['is_admin'])
                messages.success(request, f'Права администратора удалены у пользователя {user.username}')
            elif action == 'delete':
                username = user.username
                user.delete()
                messages.success(request, f'Пользователь {username} удален')

    # Оптимизированный запрос пользователей
    users = CustomUser.objects.select_related('created_by').order_by('username')
    return render(request, 'calendar_app/user_management.html', {'users': users})


@user_passes_test(is_admin)
def create_user(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.created_by = request.user
            user.save()
            messages.success(request, f'Пользователь {user.username} успешно создан!')
            return redirect('user_management')
    else:
        form = UserRegistrationForm()

    return render(request, 'calendar_app/create_user.html', {'form': form})


@user_passes_test(is_admin)
def export_selections(request):
    # Оптимизированный запрос с select_related
    selections = UserSelection.objects.select_related(
        'user', 'day_menu', 'selected_salad', 'selected_soup',
        'selected_main', 'selected_side', 'selected_bakery'
    ).all()

    data = []

    for selection in selections:
        data.append({
            'User': selection.user.username,
            'Date': selection.day_menu.date,
            'Day': selection.day_menu.get_day_display(),
            'Selected Salad': selection.selected_salad.name if selection.selected_salad else 'Не выбрано',
            'Selected Soup': selection.selected_soup.name if selection.selected_soup else 'Не выбрано',
            'Selected Main Course': selection.selected_main.name if selection.selected_main else 'Не выбрано',
            'Selected Side': selection.selected_side.name if selection.selected_side else 'Не выбрано',
            'Selected Bakery': selection.selected_bakery.name if selection.selected_bakery else 'Не выбрано',
            'Not Eating': 'Да' if selection.not_eating else 'Нет',
        })

    df = pd.DataFrame(data)
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="meal_selections.xlsx"'
    df.to_excel(response, index=False)
    return response


@user_passes_test(is_admin)
def clear_calendar(request):
    if request.method == 'POST':
        try:
            with transaction.atomic():
                UserSelection.objects.all().delete()
                DayMenu.objects.all().delete()
                Meal.objects.all().delete()
            logger.info('Calendar, dishes and user selections cleared successfully')
            messages.success(request, 'Календарь, блюда и выборы пользователей успешно очищены')
        except Exception as e:
            logger.error(f'Error clearing calendar: {str(e)}')
            messages.error(request, f'Ошибка при очистке: {str(e)}')
    return redirect('home')


@login_required
def change_password(request, user_id=None):
    # Если user_id указан и пользователь админ, меняем пароль другого пользователя
    if user_id and (request.user.is_admin or request.user.is_superuser):
        target_user = get_object_or_404(CustomUser, id=user_id)
        if request.method == 'POST':
            form = SetPasswordForm(target_user, request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, f'Пароль пользователя {target_user.username} успешно изменен')
                return redirect('user_management')
        else:
            form = SetPasswordForm(target_user)
            # Перевод меток полей на русский
            form.fields['new_password1'].label = 'Новый пароль'
            form.fields['new_password2'].label = 'Подтверждение нового пароля'
            form.fields['new_password1'].help_text = ''
            form.fields['new_password2'].help_text = ''
        return render(request, 'calendar_app/change_password.html', {
            'form': form,
            'target_user': target_user,
            'is_own_password': False
        })

    # Иначе меняем свой пароль
    else:
        if request.method == 'POST':
            form = PasswordChangeForm(request.user, request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, 'Ваш пароль успешно изменен')
                return redirect('home')
        else:
            form = PasswordChangeForm(request.user)
            # Перевод меток полей на русский
            form.fields['old_password'].label = 'Текущий пароль'
            form.fields['new_password1'].label = 'Новый пароль'
            form.fields['new_password2'].label = 'Подтверждение нового пароля'
            form.fields['old_password'].help_text = ''
            form.fields['new_password1'].help_text = ''
            form.fields['new_password2'].help_text = ''
        return render(request, 'calendar_app/change_password.html', {
            'form': form,
            'target_user': request.user,
            'is_own_password': True
        })


def get_category_rows():
    """Возвращает диапазоны строк для каждой категории"""
    return {
        'Салаты': (3, 6),  # с 3 по 6 строку
        'Супы': (7, 8),  # с 7 по 8 строку
        'Горячие блюда': (9, 14),  # с 9 по 14 строку
        'Гарниры': (15, 19),  # с 15 по 19 строку
        'Выпечка': (20, 23)  # с 20 по 23 строку
    }


@login_required
def export_data(request):
    logger = logging.getLogger(__name__)

    # Добавляем timeout для предотвращения зависания (только для Unix-систем)
    def timeout_handler(signum, frame):
        logger.error("Export operation timed out after 5 minutes")
        raise TimeoutError("Export operation timed out")

    # Устанавливаем timeout на 5 минут (только для Unix-систем)
    if hasattr(signal, 'SIGALRM'):
        signal.signal(signal.SIGALRM, timeout_handler)
        signal.alarm(300)

    try:
        # Получаем даты для следующей недели
        today = timezone.now().date()
        next_week_start = today + timedelta(days=(7 - today.weekday()))
        next_week_end = next_week_start + timedelta(days=4)

        logger.info(f"Готовим данные для экспорта: {next_week_start} - {next_week_end}")

        # Получаем меню на следующую неделю с оптимизированным запросом
        menus = DayMenu.objects.filter(
            date__range=[next_week_start, next_week_end]
        ).order_by('date')

        logger.info(f"Найдено меню: {menus.count()} дней")

        if not menus.exists():
            messages.error(request, "Меню на следующую неделю не найдено")
            return redirect('home')

        # Ищем файл меню в директории с улучшенной обработкой ошибок
        try:
            menu_files = [f for f in os.listdir(settings.MEDIA_ROOT)
                          if f.startswith('menu_') and f.endswith('.xlsx')]
        except OSError as e:
            logger.error(f"Ошибка доступа к директории media: {e}")
            messages.error(request, "Ошибка доступа к файлам меню")
            return redirect('home')

        if not menu_files:
            messages.error(request, "Файл меню не найден")
            return redirect('home')

        # Берем самый последний файл
        latest_file = sorted(menu_files)[-1]
        menu_file_path = os.path.join(settings.MEDIA_ROOT, latest_file)
        logger.info(f"Используем файл меню: {latest_file}")

        # Проверяем доступность файла
        if not os.path.exists(menu_file_path):
            logger.error(f"Файл меню не существует: {menu_file_path}")
            messages.error(request, "Файл меню недоступен")
            return redirect('home')

        # Открываем существующий файл с улучшенной обработкой ошибок
        try:
            wb = load_workbook(menu_file_path)
            ws = wb.active
        except Exception as e:
            logger.error(f"Ошибка открытия Excel файла: {e}")
            messages.error(request, "Ошибка при открытии файла меню")
            return redirect('home')

        # Маппинг дней недели к колонкам Excel (для блюд и для подсчета)
        day_columns = {
            0: ('B', 'C'),  # Понедельник (блюда в B, подсчет в C)
            1: ('D', 'E'),  # Вторник (блюда в D, подсчет в E)
            2: ('F', 'G'),  # Среда (блюда в F, подсчет в G)
            3: ('H', 'I'),  # Четверг (блюда в H, подсчет в I)
            4: ('J', 'K')  # Пятница (блюда в J, подсчет в K)
        }

        # Для каждого дня недели
        for day_offset in range(5):  # Пн-Пт
            current_date = next_week_start + timedelta(days=day_offset)
            meal_column, count_column = day_columns[day_offset]

            logger.info(f"\nОбработка дня: {current_date} (колонки {meal_column}/{count_column})")

            # Получаем меню на этот день
            day_menu = menus.filter(date=current_date).first()
            if not day_menu:
                logger.info(f"Не найдено меню на {current_date}")
                continue

            # Получаем все выборы на этот день с оптимизированным запросом
            selections = UserSelection.objects.filter(
                day_menu=day_menu,
                not_eating=False
            ).select_related('selected_salad', 'selected_soup', 'selected_main',
                             'selected_side', 'selected_bakery')

            logger.info(f"Найдено {selections.count()} выборов на {current_date}")

            # Инициализируем словарь для подсчета выборов
            meal_counts = {}

            # Подсчитываем выборы для каждого блюда
            for selection in selections:
                # Проверяем каждый тип блюда
                for meal_field in ['selected_salad', 'selected_soup', 'selected_main',
                                   'selected_side', 'selected_bakery']:
                    meal = getattr(selection, meal_field)
                    if meal and meal.excel_row:
                        if meal.excel_row not in meal_counts:
                            meal_counts[meal.excel_row] = 0
                        meal_counts[meal.excel_row] += 1

            # Записываем результаты в Excel с обработкой ошибок
            for row, count in meal_counts.items():
                try:
                    cell = ws[f'{count_column}{row}']
                    cell.value = count
                except Exception as e:
                    logger.warning(f"Ошибка записи в ячейку {count_column}{row}: {e}")

        # Сохраняем изменения в тот же файл с резервным копированием
        try:
            # Создаем резервную копию
            backup_path = menu_file_path + '.backup'
            shutil.copy2(menu_file_path, backup_path)

            # Сохраняем изменения
            wb.save(menu_file_path)
            logger.info("Файл успешно сохранен")

            # Удаляем резервную копию после успешного сохранения
            if os.path.exists(backup_path):
                os.remove(backup_path)

        except Exception as e:
            logger.error(f"Ошибка сохранения файла: {e}")
            # Восстанавливаем из резервной копии
            backup_path = menu_file_path + '.backup'
            if os.path.exists(backup_path):
                shutil.move(backup_path, menu_file_path)
                logger.info("Файл восстановлен из резервной копии")
            messages.error(request, "Ошибка при сохранении файла экспорта")
            return redirect('home')
        finally:
            wb.close()

        # Отправляем файл как ответ
        try:
            with open(menu_file_path, 'rb') as f:
                response = HttpResponse(
                    f.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response[
                    'Content-Disposition'] = f'attachment; filename=menu_with_selections_{next_week_start.strftime("%Y%m%d")}.xlsx'

            return response
        except Exception as e:
            logger.error(f"Ошибка отправки файла: {e}")
            messages.error(request, "Ошибка при отправке файла")
            return redirect('home')

    except TimeoutError:
        messages.error(request, "Операция экспорта превысила время ожидания. Попробуйте позже.")
        return redirect('home')
    except Exception as e:
        logger.error(f"Ошибка при экспорте: {str(e)}", exc_info=True)
        messages.error(request, f"Произошла ошибка при экспорте данных: {str(e)}")
        return redirect('home')
    finally:
        # Отменяем timeout (только для Unix-систем)
        if hasattr(signal, 'SIGALRM'):
            signal.alarm(0)


def parse_meal_name(cell_value):
    """Простая и безопасная версия парсинга названий блюд"""
    if not cell_value or not isinstance(cell_value, str):
        return None, None

    cell_value = cell_value.strip()
    if not cell_value:
        return None, None

    # Простое разделение по скобкам
    if '(' in cell_value and ')' in cell_value:
        try:
            parts = cell_value.split('(')
            if len(parts) >= 2:
                name = parts[0].strip()
                description = parts[1].split(')')[0].strip()
                return name, description if description else None
        except:
            pass

    return cell_value, None


def parse_excel_smart(file_path, next_week_start):
    """
    Smart parser that handles meal descriptions and Excel coordinates
    """
    logger.info("Starting smart parser")
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    # Маппинг дней к колонкам Excel (только колонки с блюдами)
    day_columns = {
        0: 'B',  # Понедельник
        1: 'D',  # Вторник
        2: 'F',  # Среда
        3: 'H',  # Четверг
        4: 'J'  # Пятница
    }

    # Создаем категории если их нет
    categories = {}
    for category_name in ['Салаты', 'Супы', 'Горячие блюда', 'Гарниры', 'Выпечка']:
        category, _ = FoodCategory.objects.get_or_create(name=category_name)
        categories[category_name] = category

    # Определяем диапазоны строк для категорий, читая из колонки A
    category_ranges = {}
    current_category = None
    start_row = None

    # Сканируем колонку A для определения диапазонов категорий
    for row in range(3, ws.max_row + 1):
        cell_value = ws['A{}'.format(row)].value
        if cell_value:
            cell_value = str(cell_value).strip().lower()
            # Если нашли категорию
            for category_name in categories.keys():
                if category_name.lower() in cell_value:
                    # Если была предыдущая категория, сохраняем её диапазон
                    if current_category and start_row:
                        category_ranges[current_category] = (start_row, row - 1)
                    # Начинаем новую категорию
                    current_category = category_name
                    start_row = row
                    break

    # Добавляем последнюю категорию
    if current_category and start_row:
        category_ranges[current_category] = (start_row, ws.max_row)

    logger.info("Category ranges found:")
    for category, (start, end) in category_ranges.items():
        logger.info(f"{category}: rows {start}-{end}")

    # Для каждого дня недели
    for day_idx, column in day_columns.items():
        day_date = next_week_start + timedelta(days=day_idx)
        day_menu = DayMenu.objects.create(date=day_date)
        logger.info(f"\nProcessing menu for {day_date} (column {column})")

        # Обрабатываем каждую категорию
        for category_name, (start_row, end_row) in category_ranges.items():
            category = categories[category_name]
            logger.info(f"\nProcessing {category_name} (rows {start_row}-{end_row})")

            # Читаем блюда для этой категории
            for current_row in range(start_row, end_row + 1):
                cell = ws[f'{column}{current_row}']
                if cell.value:
                    # Парсим название и описание блюда
                    name, description = parse_meal_name(str(cell.value))
                    if name:
                        try:
                            # Всегда создаем новое блюдо
                            meal = Meal.objects.create(
                                name=name,
                                category=category,
                                description=description,
                                excel_row=current_row
                            )

                            # Добавляем блюдо в соответствующую категорию меню
                            if category_name == 'Салаты':
                                day_menu.salads.add(meal)
                            elif category_name == 'Супы':
                                day_menu.soups.add(meal)
                            elif category_name == 'Горячие блюда':
                                day_menu.main_courses.add(meal)
                            elif category_name == 'Гарниры':
                                day_menu.sides.add(meal)
                            elif category_name == 'Выпечка':
                                day_menu.bakery.add(meal)

                            logger.info(
                                f"Created new meal: {name} in {category_name} for {day_date} (row {current_row})")

                        except Exception as e:
                            logger.error(f"Error adding meal '{name}': {str(e)}")

    wb.close()
    logger.info("Smart parser completed successfully")


def is_admin_or_root(user):
    return user.is_superuser or user.is_staff or user.is_admin


@user_passes_test(is_admin_or_root)
def manage_dishes(request):
    # Get only the "Горячие блюда" category with optimized query
    try:
        hot_dishes_category = FoodCategory.objects.get(name='Горячие блюда')
        # Get dishes for hot dishes category
        dishes = Meal.objects.filter(category=hot_dishes_category).order_by('name')
    except FoodCategory.DoesNotExist:
        dishes = []
        messages.warning(request, 'Категория "Горячие блюда" не найдена')

    return render(request, 'calendar_app/manage_dishes.html', {
        'dishes': dishes
    })


@user_passes_test(is_admin_or_root)
def update_complete_dish_status(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Метод не поддерживается'}, status=405)

    try:
        data = json.loads(request.body)
        meal_id = data.get('meal_id')
        is_complete = data.get('is_complete', False)

        # Используем select_for_update для предотвращения race conditions
        with transaction.atomic():
            meal = Meal.objects.select_for_update().get(id=meal_id)

            # Проверяем, что блюдо относится к категории "Горячие блюда"
            if meal.category.name != 'Горячие блюда':
                return JsonResponse({
                    'error': 'Только горячие блюда могут быть отмечены как полноценные'
                }, status=400)

            meal.is_complete_dish = is_complete
            meal.save(update_fields=['is_complete_dish'])

        return JsonResponse({
            'success': True,
            'message': f'Статус блюда "{meal.name}" успешно обновлен',
            'is_complete': meal.is_complete_dish
        })

    except Meal.DoesNotExist:
        return JsonResponse({'error': 'Блюдо не найдено'}, status=404)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Неверный формат данных'}, status=400)
    except Exception as e:
        logger.error(f'Error updating complete dish status: {str(e)}')
        return JsonResponse({'error': 'Произошла ошибка при обновлении статуса'}, status=500)


@user_passes_test(is_admin_or_root)
def clear_all_dishes(request):
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Delete all meals from the database
                meal_count = Meal.objects.count()
                Meal.objects.all().delete()
                logger.info(f'Cleared {meal_count} dishes from database')
                messages.success(request, f'Все блюда ({meal_count}) успешно удалены')
        except Exception as e:
            logger.error(f'Error clearing all dishes: {str(e)}')
            messages.error(request, f'Произошла ошибка при удалении блюд: {str(e)}')

    return redirect('manage_dishes')